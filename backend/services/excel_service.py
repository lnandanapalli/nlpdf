"""Excel (XLSX) to PDF conversion service using openpyxl and xhtml2pdf."""

import structlog
from html import escape
from pathlib import Path

from openpyxl import load_workbook
from xhtml2pdf import pisa

from backend.services.pdf_fonts import register_fonts

logger = structlog.get_logger(__name__)

CSS = """\
body {
    font-family: DejaVuSans, Helvetica, Arial, sans-serif;
    font-size: 11px;
    line-height: 1.4;
    color: #222;
    margin: 30px;
}
h2 {
    font-size: 18px;
    margin-top: 24px;
    margin-bottom: 10px;
    color: #1a1a2e;
}
table {
    border-collapse: collapse;
    width: 100%;
    margin-bottom: 20px;
}
th, td {
    border: 1px solid #ccc;
    padding: 6px 10px;
    text-align: left;
    word-wrap: break-word;
}
th {
    background-color: #e8e8e8;
    font-weight: bold;
}
tr:nth-child(even) td {
    background-color: #f9f9f9;
}
.sheet-separator {
    page-break-before: always;
}
"""

PAPER_SIZES = {
    "A4": "@page { size: A4 landscape; margin: 1.5cm; }",
    "letter": "@page { size: letter landscape; margin: 0.75in; }",
}


def _sheet_to_html(ws) -> str:
    """Convert a worksheet to an HTML table."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""

    # Skip fully empty rows
    non_empty = [r for r in rows if any(c is not None for c in r)]
    if not non_empty:
        return ""

    html_rows = []
    for i, row in enumerate(non_empty):
        tag = "th" if i == 0 else "td"
        cells = "".join(
            f"<{tag}>{escape(str(cell)) if cell is not None else ''}</{tag}>"
            for cell in row
        )
        html_rows.append(f"<tr>{cells}</tr>")

    return f"<table>{''.join(html_rows)}</table>"


def excel_to_pdf(input_path: Path, output_path: Path, paper_size: str = "A4") -> Path:
    """
    Convert an Excel (XLSX) file to PDF.

    Reads all sheets with openpyxl and renders each as an HTML table,
    then converts to PDF via xhtml2pdf.

    Args:
        input_path: Path to the input XLSX file.
        output_path: Path to save the output PDF.
        paper_size: Paper size — "A4" or "letter".

    Returns:
        Path to the generated PDF file.

    Raises:
        ValueError: If the XLSX file cannot be read or converted.
    """
    register_fonts()

    wb = load_workbook(str(input_path), read_only=True, data_only=True)

    sheets_html = []
    for i, name in enumerate(wb.sheetnames):
        ws = wb[name]
        table_html = _sheet_to_html(ws)
        if not table_html:
            continue

        css_class = ' class="sheet-separator"' if i > 0 else ""
        sheets_html.append(
            f"<div{css_class}>" f"<h2>{escape(name)}</h2>" f"{table_html}" f"</div>"
        )

    wb.close()

    body = "\n".join(sheets_html) if sheets_html else "<p>Empty spreadsheet</p>"
    page_css = PAPER_SIZES.get(paper_size, PAPER_SIZES["A4"])

    html = (
        "<!DOCTYPE html>"
        "<html><head><meta charset='utf-8'/>"
        f"<style>{page_css}\n{CSS}</style>"
        f"</head><body>{body}</body></html>"
    )

    with open(output_path, "wb") as f:
        status = pisa.CreatePDF(html, dest=f)

    if status.err:
        output_path.unlink(missing_ok=True)
        raise ValueError(f"PDF conversion failed with {status.err} error(s)")

    logger.info(
        "Excel converted to PDF",
        input=str(input_path),
        output=str(output_path),
        sheets=len(sheets_html),
        paper_size=paper_size,
    )

    return output_path
