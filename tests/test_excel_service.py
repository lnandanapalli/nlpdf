"""Tests for Excel (XLSX) to PDF conversion service."""

from pathlib import Path

import pytest
from openpyxl import Workbook

from backend.services.excel_service import excel_to_pdf


@pytest.fixture()
def xlsx_file(tmp_path: Path) -> Path:
    """Create a sample XLSX file with data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 30, "NYC"])
    ws.append(["Bob", 25, "London"])
    ws.append(["Charlie", 35, "Tokyo"])
    path = tmp_path / "sample.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture()
def xlsx_multi_sheet(tmp_path: Path) -> Path:
    """Create an XLSX file with multiple sheets."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["Product", "Revenue"])
    ws1.append(["Widget", 1000])
    ws1.append(["Gadget", 2500])

    ws2 = wb.create_sheet("Expenses")
    ws2.append(["Category", "Amount"])
    ws2.append(["Rent", 500])
    ws2.append(["Utilities", 200])

    path = tmp_path / "multi.xlsx"
    wb.save(str(path))
    return path


class TestExcelToPdf:
    """Tests for excel_to_pdf."""

    def test_basic_conversion(self, xlsx_file: Path, tmp_path: Path):
        output = tmp_path / "output.pdf"
        result = excel_to_pdf(xlsx_file, output)

        assert result == output
        assert output.exists()
        assert output.stat().st_size > 0
        assert output.read_bytes()[:5] == b"%PDF-"

    def test_multi_sheet_conversion(self, xlsx_multi_sheet: Path, tmp_path: Path):
        output = tmp_path / "output.pdf"
        result = excel_to_pdf(xlsx_multi_sheet, output)

        assert result == output
        assert output.exists()
        assert output.read_bytes()[:5] == b"%PDF-"

    def test_letter_paper_size(self, xlsx_file: Path, tmp_path: Path):
        output = tmp_path / "output.pdf"
        result = excel_to_pdf(xlsx_file, output, paper_size="letter")

        assert result == output
        assert output.exists()
        assert output.read_bytes()[:5] == b"%PDF-"

    def test_empty_xlsx(self, tmp_path: Path):
        wb = Workbook()
        path = tmp_path / "empty.xlsx"
        wb.save(str(path))
        output = tmp_path / "output.pdf"

        result = excel_to_pdf(path, output)
        assert result == output
        assert output.exists()
